#!/usr/bin/python
# -*- coding:utf-8 -*-
'''

@author: fengjj
'''

import urllib, urllib2, sys
import ssl
import json
from openpyxl import Workbook
from openpyxl import load_workbook


#公司编码，根据订单号，查询订单状态，并返回
def getNewStatus(comp ,orderNo):
    host = 'https://ali-deliver.showapi.com'
    path = '/showapi_expInfo'
    method = 'GET'
    appcode = 'ede19edebccb424a8ded61ad3ae52f9c'
    querys = 'com='+comp+'&nu='+orderNo
    bodys = {}
    url = host + path + '?' + querys

    request = urllib2.Request(url)
    request.add_header('Authorization', 'APPCODE ' + appcode)
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    response = urllib2.urlopen(request, context=ctx)
    content = response.read()
    if (content):
        print(content)
        #status": 4,-1 待查询 0 查询异常 1 暂无记录 2 在途中 3 派送中 4 已签收 5 用户拒签 6 疑难件 7 无效单 8 超时单 9 签收失败 10 退回
        jsonContent = json.loads(content)
        return jsonContent["showapi_res_body"]["status"],jsonContent["showapi_res_body"] ["data"][0]["time"]
    else:
        return -1

def main(path,filename):
    wb = load_workbook(path+filename)
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])
    #wb = copy(data)
    nrows = ws.max_row
    print (nrows)
    for i in range(2, nrows):
        orderNo = ws["F"+str(i)].value
        comp = ws["D"+str(i)].value
        oStatus = ws["O"+str(i)].value
        if orderNo and comp and oStatus!="yes":
            print str(comp) + " " + str(orderNo) + " " + str(oStatus)
            ret = getNewStatus(str(comp), str(orderNo))
            recTime = ret[1]
            nstatus = ret[0]
            print nstatus
            if (4 == nstatus):
                # table.cell(0, 0).value = "yes"
                # table.put_cell(i, 13, 1, "success",0)
                ws["O" + str(i)].value = 'yes'
                ws["P" + str(i)].value = recTime
                wb.save(path + filename)
                print str(orderNo) + "已经改写成签收"
            else:
                ws["O" + str(i)].value = 'no'
                wb.save(path + filename)
                print str(orderNo) + "已经改写成 no 未签收 ，状态：" + str(nstatus)

if __name__ == '__main__':
    #文件默认路径
    path = 'C:\\Users\\fengjj\\Desktop\\'
    #从参数读文件名
    if len(sys.argv)<2:
        filename='1yuekuaidi.xlsx'
    else:
        filename = sys.argv[1]
    main(path,filename)
